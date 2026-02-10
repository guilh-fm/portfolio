import os
import threading
import openpyxl
from openpyxl import load_workbook
import customtkinter as ctk
from tkinter import filedialog, messagebox

class PreenchFormsOi:
    def __init__(self, master):
        self.master = master
        
        master.title("PFOI - Preenche Formulários de Interconexão OI")
        master.geometry("650x800")
        master.resizable(False, False)
        
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")
        
        self.main_frame = ctk.CTkFrame(master, fg_color="#181818", corner_radius=0)
        self.main_frame.pack(fill="both", expand=True)

        self.caminho_xlsx = ctk.StringVar()
        self.caminho_modelo = ctk.StringVar()
        self.caminho_pasta_saida = ctk.StringVar()
        self.prefixo_saida = ctk.StringVar(value="ITX-OPERADORA")

        self.frame_selecao_arquivos = ctk.CTkFrame(self.main_frame, corner_radius=10, fg_color="white", border_width=2, border_color="#383838")
        self.frame_selecao_arquivos.pack(pady=10, padx=20, fill="x")
        ctk.CTkLabel(self.frame_selecao_arquivos, text="1. Selecione os Arquivos", font=('Arial', 12, 'bold'), text_color="#181818").pack(pady=5, padx=10, anchor="w")

        ctk.CTkLabel(self.frame_selecao_arquivos, text="Base de Dados - Excel:", text_color="#181818").pack(pady=5, padx=10, anchor="w")
        frame_input_excel = ctk.CTkFrame(self.frame_selecao_arquivos, fg_color="transparent")
        frame_input_excel.pack(fill="x", padx=10, pady=(5,20))
        
        self.entrada_excel = ctk.CTkEntry(frame_input_excel, textvariable=self.caminho_xlsx, width=450, fg_color="#383838", text_color="white")
        self.entrada_excel.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.btn_excel = ctk.CTkButton(frame_input_excel, text="Procurar...", command=self.procurar_excel, fg_color="#181818", hover_color="#3B3E41", text_color="white")
        self.btn_excel.pack(side="right")

        ctk.CTkLabel(self.frame_selecao_arquivos, text="Formulário - Excel Modelo:", text_color="#181818").pack(pady=5, padx=10, anchor="w")
        frame_input_modelo = ctk.CTkFrame(self.frame_selecao_arquivos, fg_color="transparent")
        frame_input_modelo.pack(fill="x", padx=10, pady=(5,20))
        
        self.entrada_modelo = ctk.CTkEntry(frame_input_modelo, textvariable=self.caminho_modelo, width=450, fg_color="#383838", text_color="white")
        self.entrada_modelo.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.btn_modelo = ctk.CTkButton(frame_input_modelo, text="Procurar...", command=self.procurar_modelo, fg_color="#181818", hover_color="#3B3E41", text_color="white")
        self.btn_modelo.pack(side="right")

        self.frame_config_saida = ctk.CTkFrame(self.main_frame, corner_radius=10, fg_color="white", border_width=2, border_color="#383838")
        self.frame_config_saida.pack(pady=10, padx=20, fill="x")
        ctk.CTkLabel(self.frame_config_saida, text="2. Configurações de Saída", font=('Arial', 12, 'bold'), text_color="#181818").pack(pady=5, padx=10, anchor="w")

        ctk.CTkLabel(self.frame_config_saida, text="Prefixo do Nome do Arquivo de Saída:", text_color="#181818").pack(pady=5, padx=10, anchor="w")
        self.entrada_prefixo = ctk.CTkEntry(self.frame_config_saida, textvariable=self.prefixo_saida, fg_color="#383838", text_color="white")
        self.entrada_prefixo.pack(pady=5, padx=10, fill="x", expand=True)
        
        ctk.CTkLabel(self.frame_config_saida, text="Diretório para Salvar Documentos:", text_color="#181818").pack(pady=5, padx=10, anchor="w")
        frame_input_dir = ctk.CTkFrame(self.frame_config_saida, fg_color="transparent")
        frame_input_dir.pack(fill="x", padx=10, pady=(5,20))
        
        self.entrada_dir_saida = ctk.CTkEntry(frame_input_dir, textvariable=self.caminho_pasta_saida, width=450, fg_color="#383838", text_color="white")
        self.entrada_dir_saida.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.btn_dir_saida = ctk.CTkButton(frame_input_dir, text="Procurar Pasta...", command=self.procurar_pasta_saida, fg_color="#181818", hover_color="#3B3E41", text_color="white")
        self.btn_dir_saida.pack(side="right")

        self.frame_acao = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.frame_acao.pack(pady=10, padx=20)
        
        self.btn_gerar = ctk.CTkButton(self.frame_acao, text="Gerar Documentos Excel", command=self.iniciar_thread_geracao, fg_color="#383838", hover_color="#7C8B8F", text_color="white", border_width=2, border_color="white")
        self.btn_gerar.pack(pady=10, fill="x", expand=True)
        
        self.frame_status = ctk.CTkFrame(self.main_frame, corner_radius=10, fg_color="white", border_width=2, border_color="#383838")
        self.frame_status.pack(pady=10, padx=20, fill="both", expand=True)
        ctk.CTkLabel(self.frame_status, text="Status do Processo", font=('Arial', 12, 'bold'), text_color="#181818").pack(pady=5, padx=10, anchor="w")
        
        self.barra_progresso = ctk.CTkProgressBar(self.frame_status, orientation="horizontal", mode="determinate", progress_color="#383838", corner_radius=10)
        self.barra_progresso.pack(pady=5, padx=10, fill="x")
        self.barra_progresso.set(0)

        self.texto_status = ctk.CTkTextbox(self.frame_status, height=150, wrap="word", font=('Arial', 11), corner_radius=10, fg_color="#383838", text_color="#181818")
        self.texto_status.pack(pady=10, padx=10, fill="both", expand=True)
        
        self.texto_status.tag_config("sucesso", foreground="green")
        self.texto_status.tag_config("erro", foreground="red")
        self.texto_status.tag_config("info", foreground="white")
        self.texto_status.tag_config("padrao", foreground="#181818")

        self.atualizar_log("Aguardando seleção de arquivos...", "info")

    def atualizar_log(self, mensagem, tag="padrao"):
        self.texto_status.configure(state="normal")
        self.texto_status.insert("end", mensagem + "\n", tag)
        self.texto_status.see("end")
        self.texto_status.configure(state="disabled")
        self.master.update_idletasks()

    def procurar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel (XLSX) com os dados",
            filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
        )
        if arquivo:
            self.caminho_xlsx.set(arquivo)
            self.atualizar_log(f"Arquivo XLSX selecionado: {os.path.basename(arquivo)}")

    def procurar_modelo(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o template Excel (XLSX) do Formulário",
            filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
        )
        if arquivo:
            self.caminho_modelo.set(arquivo)
            self.atualizar_log(f"Template XLSX selecionado: {os.path.basename(arquivo)}")

    def procurar_pasta_saida(self):
        diretorio = filedialog.askdirectory(
            title="Selecione o diretório para salvar os documentos gerados"
        )
        if diretorio:
            self.caminho_pasta_saida.set(diretorio)
            self.atualizar_log(f"Diretório de saída selecionado: {os.path.basename(diretorio)}")

    def iniciar_thread_geracao(self):
        thread = threading.Thread(target=self.executar_geracao)
        thread.daemon = True
        thread.start()

    def executar_geracao(self):
        arquivo_excel = self.caminho_xlsx.get()
        arquivo_modelo = self.caminho_modelo.get()
        pasta_saida = self.caminho_pasta_saida.get()
        prefixo_saida = self.prefixo_saida.get().strip()

        if not arquivo_excel or not arquivo_modelo or not pasta_saida:
            self.master.after(0, lambda: messagebox.showerror("Erro de Seleção", "Por favor, selecione todos os arquivos e diretórios."))
            return

        self.master.after(0, self.atualizar_log, "Iniciando geração de planilhas...", "info")
        self.master.after(0, lambda: self.btn_gerar.configure(state="disabled"))
        
        try:
            dir_destino = os.path.join(pasta_saida, "FORMULARIOS_PREENCHIDOS_XLSX")
            if not os.path.exists(dir_destino):
                os.makedirs(dir_destino)

            self.preencher_planilhas_xlsx(arquivo_excel, arquivo_modelo, dir_destino, prefixo_saida)
            
            self.master.after(0, self.atualizar_log, "\nGerados todos os arquivos XLSX com sucesso.", "sucesso")
            self.master.after(0, lambda: messagebox.showinfo("Sucesso", f"Planilhas XLSX geradas com sucesso em:\n{dir_destino}"))

        except Exception as e:
            self.master.after(0, self.atualizar_log, f"\nOcorreu um erro: {e}", "erro")
            self.master.after(0, lambda: messagebox.showerror("Erro", f"Ocorreu um erro durante a geração:\n{e}"))
        finally:
            self.master.after(0, lambda: self.btn_gerar.configure(state="normal"))

    def preencher_planilhas_xlsx(self, caminho_planilha_xlsx, caminho_template_xlsx, diretorio_saida, prefixo_base):
        try:
            wb_base = load_workbook(filename=caminho_planilha_xlsx, data_only=True)
            aba_base = wb_base.active
            total_linhas = aba_base.max_row - 2
            
            if total_linhas <= 0:
                raise Exception("A planilha de dados está vazia ou não tem dados após o cabeçalho.")
                
            self.master.after(0, lambda: self.barra_progresso.set(0))

            mapeamento_colunas = {
                "CIDADE": 0,
                "ESTADO": 2,
                "CN": 3,
                "AREA_LOCAL": 4,
                "SIGLA": 5,
                "PREFIXO": 9,
                "INICIAL": 10,
                "FINAL": 11,
                "EOT": 13,
                "RN1": 15,
                "ENDERECO": 18,
                "CEP": 19,
                "LATITUDE": 20,
                "LONGITUDE": 21,
            }
            
            indice_uf = 1
            indice_area_local = 4

        except Exception as e:
            raise Exception(f"Erro ao carregar a planilha XLSX de dados: {e}")

        for i, dados_linha in enumerate(aba_base.iter_rows(min_row=3, values_only=True)):
            if not any(dados_linha):
                self.master.after(0, self.atualizar_log, f"  Pulando linha {i + 3} (vazia)...", "info")
                total_linhas -= 1 
                continue

            self.master.after(0, self.atualizar_log, f"  Processando linha {i + 3}...", "info")
            
            try:
                wb_modelo = load_workbook(filename=caminho_template_xlsx)
            except Exception as e:
                raise Exception(f"Erro ao carregar o template XLSX: {e}")

            valores_substituir = {}
            for marcador, col_index in mapeamento_colunas.items():
                valor = dados_linha[col_index] if col_index < len(dados_linha) else ""
                valores_substituir[marcador] = str(valor) if valor is not None else ""

            for aba_modelo in wb_modelo.worksheets:
                for linha_modelo in aba_modelo.iter_rows():
                    for celula in linha_modelo:
                        if isinstance(celula.value, str) and "{{" in celula.value and "}}" in celula.value:
                            novo_valor = celula.value
                            alterado = False
                            
                            for codigo, valor_real in valores_substituir.items():
                                tag = f"{{{{{codigo}}}}}"
                                if tag in novo_valor:
                                    novo_valor = novo_valor.replace(tag, valor_real)
                                    alterado = True
                            
                            if alterado:
                                celula.value = novo_valor

            try:
                aba_form = wb_modelo["Formulário"]
                valor_uf = str(dados_linha[indice_uf]) if indice_uf < len(dados_linha) and dados_linha[indice_uf] is not None else ""
                aba_form["C37"] = valor_uf
            except KeyError:
                self.master.after(0, self.atualizar_log, "  AVISO: Aba 'Formulário' não encontrada. Pulando regra C37.", "erro")
            except Exception as e:
                self.master.after(0, self.atualizar_log, f"  AVISO: Erro ao preencher C37: {e}", "erro")
            
            nome_area = str(dados_linha[indice_area_local]) if indice_area_local < len(dados_linha) and dados_linha[indice_area_local] is not None else "Sem_Area_Local"
            
            nome_arquivo = f"{prefixo_base}-OI-{nome_area}.xlsx"
            caminho_final = os.path.join(diretorio_saida, nome_arquivo)

            try:
                wb_modelo.save(caminho_final)
                self.master.after(0, self.atualizar_log, f"  Gerado: {os.path.basename(caminho_final)}")
                
                if total_linhas > 0:
                    progresso = (i + 1) / total_linhas
                    self.master.after(0, lambda p=progresso: self.barra_progresso.set(p))
            
            except Exception as e:
                self.master.after(0, self.atualizar_log, f"Erro ao salvar '{os.path.basename(caminho_final)}': {e}", "erro")

if __name__ == "__main__":
    raiz = ctk.CTk()
    app = PreenchFormsOi(raiz)
    raiz.mainloop()