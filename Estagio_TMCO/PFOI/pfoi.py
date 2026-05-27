import os
import threading
from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import load_workbook


MAPEAMENTO_COLUNAS = {
    "CIDADE": 0,
    "ESTADO": 2,
    "CN": 3,
    "AREA_LOCAL": 4,
    "SIGLA": 5,
    "PREFIXO": 9,
    "INICIAL": 10,
    "FINAL": 11,
    "EOT": 13,
    "RN1": 15,
    "ENDERECO": 18,
    "CEP": 19,
    "LATITUDE": 20,
    "LONGITUDE": 21,
}

INDICE_UF = 1
INDICE_AREA_LOCAL = 4


# Funções auxiliares

def criar_pasta_saida(pasta_base):
    pasta_destino = os.path.join(pasta_base, "FORMULARIOS_PREENCHIDOS_XLSX")

    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)

    return pasta_destino


def validar_campos_obrigatorios(caminho_excel, caminho_modelo, pasta_saida):
    if caminho_excel == "" or caminho_modelo == "" or pasta_saida == "":
        return False

    return True


def carregar_planilha(caminho_planilha):
    try:
        pasta_trabalho = load_workbook(filename=caminho_planilha, data_only=True)
        aba = pasta_trabalho.active
        return aba
    except Exception as erro:
        raise Exception(f"Erro ao carregar a planilha de dados: {erro}")


def carregar_modelo(caminho_modelo):
    try:
        modelo = load_workbook(filename=caminho_modelo)
        return modelo
    except Exception as erro:
        raise Exception(f"Erro ao carregar o formulário modelo: {erro}")


def montar_valores_substituicao(dados_linha):
    valores_substituir = {}

    for marcador, indice_coluna in MAPEAMENTO_COLUNAS.items():
        if indice_coluna < len(dados_linha) and dados_linha[indice_coluna] is not None:
            valores_substituir[marcador] = str(dados_linha[indice_coluna])
        else:
            valores_substituir[marcador] = ""

    return valores_substituir


def substituir_marcadores_na_planilha(modelo, valores_substituir):
    for aba_modelo in modelo.worksheets:
        for linha in aba_modelo.iter_rows():
            for celula in linha:
                if isinstance(celula.value, str):
                    novo_valor = celula.value

                    for marcador, valor_real in valores_substituir.items():
                        tag = "{{" + marcador + "}}"
                        novo_valor = novo_valor.replace(tag, valor_real)

                    celula.value = novo_valor


def preencher_campo_especifico(modelo, dados_linha):
    try:
        aba_formulario = modelo["Formulário"]

        if INDICE_UF < len(dados_linha) and dados_linha[INDICE_UF] is not None:
            valor_uf = str(dados_linha[INDICE_UF])
        else:
            valor_uf = ""

        aba_formulario["C37"] = valor_uf
    except KeyError:
        # Alguns modelos podem não ter essa aba. Nesse caso, o restante ainda pode ser gerado.
        pass


def montar_nome_arquivo(prefixo_base, dados_linha):
    if INDICE_AREA_LOCAL < len(dados_linha) and dados_linha[INDICE_AREA_LOCAL] is not None:
        nome_area = str(dados_linha[INDICE_AREA_LOCAL])
    else:
        nome_area = "Sem_Area_Local"

    nome_arquivo = f"{prefixo_base}-OI-{nome_area}.xlsx"
    return nome_arquivo


def salvar_modelo_preenchido(modelo, diretorio_saida, nome_arquivo):
    caminho_final = os.path.join(diretorio_saida, nome_arquivo)
    modelo.save(caminho_final)
    return caminho_final


# Interface gráfica

class AplicacaoPfoi:
    def __init__(self, janela):
        self.janela = janela

        self.janela.title("PFOI - Preenche Formulários de Interconexão OI")
        self.janela.geometry("650x800")
        self.janela.resizable(False, False)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

        self.caminho_excel = ctk.StringVar()
        self.caminho_modelo = ctk.StringVar()
        self.caminho_pasta_saida = ctk.StringVar()
        self.prefixo_saida = ctk.StringVar(value="ITX-OPERADORA")

        self.criar_componentes()
        self.atualizar_log("Aguardando seleção de arquivos...", "info")

    def criar_componentes(self):
        self.frame_principal = ctk.CTkFrame(self.janela, fg_color="#181818", corner_radius=0)
        self.frame_principal.pack(fill="both", expand=True)

        self.criar_bloco_arquivos()
        self.criar_bloco_saida()
        self.criar_bloco_acoes()
        self.criar_bloco_status()

    def criar_bloco_arquivos(self):
        frame = ctk.CTkFrame(self.frame_principal, corner_radius=10, fg_color="white", border_width=2, border_color="#383838")
        frame.pack(pady=10, padx=20, fill="x")

        ctk.CTkLabel(frame, text="1. Selecione os Arquivos", font=("Arial", 12, "bold"), text_color="#181818").pack(pady=5, padx=10, anchor="w")
        self.criar_linha_arquivo(frame, "Base de Dados - Excel:", self.caminho_excel, self.procurar_excel)
        self.criar_linha_arquivo(frame, "Formulário - Excel Modelo:", self.caminho_modelo, self.procurar_modelo)

    def criar_bloco_saida(self):
        frame = ctk.CTkFrame(self.frame_principal, corner_radius=10, fg_color="white", border_width=2, border_color="#383838")
        frame.pack(pady=10, padx=20, fill="x")

        ctk.CTkLabel(frame, text="2. Configurações de Saída", font=("Arial", 12, "bold"), text_color="#181818").pack(pady=5, padx=10, anchor="w")
        ctk.CTkLabel(frame, text="Prefixo do Nome do Arquivo de Saída:", text_color="#181818").pack(pady=5, padx=10, anchor="w")

        entrada_prefixo = ctk.CTkEntry(frame, textvariable=self.prefixo_saida, fg_color="#383838", text_color="white")
        entrada_prefixo.pack(pady=5, padx=10, fill="x")

        self.criar_linha_arquivo(frame, "Diretório para Salvar Documentos:", self.caminho_pasta_saida, self.procurar_pasta_saida, texto_botao="Procurar Pasta...")

    def criar_bloco_acoes(self):
        frame = ctk.CTkFrame(self.frame_principal, fg_color="transparent")
        frame.pack(pady=10, padx=20)

        self.botao_gerar = ctk.CTkButton(
            frame,
            text="Gerar Documentos Excel",
            command=self.iniciar_thread_geracao,
            fg_color="#383838",
            hover_color="#7C8B8F",
            text_color="white",
            border_width=2,
            border_color="white",
        )
        self.botao_gerar.pack(pady=10, fill="x")

    def criar_bloco_status(self):
        frame = ctk.CTkFrame(self.frame_principal, corner_radius=10, fg_color="white", border_width=2, border_color="#383838")
        frame.pack(pady=10, padx=20, fill="both", expand=True)

        ctk.CTkLabel(frame, text="Status do Processo", font=("Arial", 12, "bold"), text_color="#181818").pack(pady=5, padx=10, anchor="w")

        self.barra_progresso = ctk.CTkProgressBar(frame, orientation="horizontal", mode="determinate", progress_color="#383838", corner_radius=10)
        self.barra_progresso.pack(pady=5, padx=10, fill="x")
        self.barra_progresso.set(0)

        self.texto_status = ctk.CTkTextbox(frame, height=150, wrap="word", font=("Arial", 11), corner_radius=10, fg_color="#383838", text_color="white")
        self.texto_status.pack(pady=10, padx=10, fill="both", expand=True)

        self.texto_status.tag_config("sucesso", foreground="green")
        self.texto_status.tag_config("erro", foreground="red")
        self.texto_status.tag_config("info", foreground="white")
        self.texto_status.tag_config("padrao", foreground="white")

    def criar_linha_arquivo(self, frame_pai, texto_label, variavel, comando, texto_botao="Procurar..."):
        ctk.CTkLabel(frame_pai, text=texto_label, text_color="#181818").pack(pady=5, padx=10, anchor="w")

        frame_input = ctk.CTkFrame(frame_pai, fg_color="transparent")
        frame_input.pack(fill="x", padx=10, pady=(5, 20))

        entrada = ctk.CTkEntry(frame_input, textvariable=variavel, width=450, fg_color="#383838", text_color="white")
        entrada.pack(side="left", fill="x", expand=True, padx=(0, 5))

        botao = ctk.CTkButton(frame_input, text=texto_botao, command=comando, fg_color="#181818", hover_color="#3B3E41", text_color="white")
        botao.pack(side="right")

    def atualizar_log(self, mensagem, tag="padrao"):
        self.texto_status.configure(state="normal")
        self.texto_status.insert("end", mensagem + "\n", tag)
        self.texto_status.see("end")
        self.texto_status.configure(state="disabled")
        self.janela.update_idletasks()

    def procurar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel com os dados",
            filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")),
        )

        if arquivo:
            self.caminho_excel.set(arquivo)
            self.atualizar_log(f"Arquivo selecionado: {os.path.basename(arquivo)}")

    def procurar_modelo(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o formulário modelo em Excel",
            filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")),
        )

        if arquivo:
            self.caminho_modelo.set(arquivo)
            self.atualizar_log(f"Modelo selecionado: {os.path.basename(arquivo)}")

    def procurar_pasta_saida(self):
        diretorio = filedialog.askdirectory(title="Selecione o diretório para salvar os documentos")

        if diretorio:
            self.caminho_pasta_saida.set(diretorio)
            self.atualizar_log(f"Diretório de saída selecionado: {os.path.basename(diretorio)}")

    def iniciar_thread_geracao(self):
        thread = threading.Thread(target=self.executar_geracao)
        thread.daemon = True
        thread.start()

    def executar_geracao(self):
        caminho_excel = self.caminho_excel.get()
        caminho_modelo = self.caminho_modelo.get()
        pasta_saida = self.caminho_pasta_saida.get()
        prefixo_saida = self.prefixo_saida.get().strip()

        if not validar_campos_obrigatorios(caminho_excel, caminho_modelo, pasta_saida):
            self.janela.after(0, lambda: messagebox.showerror("Erro de Seleção", "Por favor, selecione todos os arquivos e diretórios."))
            return

        self.janela.after(0, self.atualizar_log, "Iniciando geração de planilhas...", "info")
        self.janela.after(0, lambda: self.botao_gerar.configure(state="disabled"))

        try:
            diretorio_saida = criar_pasta_saida(pasta_saida)
            self.preencher_planilhas(caminho_excel, caminho_modelo, diretorio_saida, prefixo_saida)

            self.janela.after(0, self.atualizar_log, "\nTodos os arquivos XLSX foram gerados com sucesso.", "sucesso")
            self.janela.after(0, lambda: messagebox.showinfo("Sucesso", f"Planilhas geradas em:\n{diretorio_saida}"))
        except Exception as erro:
            mensagem_erro = str(erro)
            self.janela.after(0, self.atualizar_log, f"\nOcorreu um erro: {mensagem_erro}", "erro")
            self.janela.after(0, lambda: messagebox.showerror("Erro", f"Ocorreu um erro durante a geração:\n{mensagem_erro}"))
        finally:
            self.janela.after(0, lambda: self.botao_gerar.configure(state="normal"))

    def preencher_planilhas(self, caminho_planilha, caminho_modelo, diretorio_saida, prefixo_saida):
        aba_dados = carregar_planilha(caminho_planilha)
        linhas_dados = list(aba_dados.iter_rows(min_row=3, values_only=True))
        linhas_validas = [linha for linha in linhas_dados if any(linha)]

        if len(linhas_validas) == 0:
            raise Exception("A planilha de dados está vazia ou não possui dados após o cabeçalho.")

        self.janela.after(0, lambda: self.barra_progresso.set(0))

        for indice, dados_linha in enumerate(linhas_validas):
            numero_linha = indice + 3
            self.janela.after(0, self.atualizar_log, f"  Processando linha {numero_linha}...", "info")

            modelo = carregar_modelo(caminho_modelo)
            valores_substituir = montar_valores_substituicao(dados_linha)

            substituir_marcadores_na_planilha(modelo, valores_substituir)
            preencher_campo_especifico(modelo, dados_linha)

            nome_arquivo = montar_nome_arquivo(prefixo_saida, dados_linha)
            caminho_final = salvar_modelo_preenchido(modelo, diretorio_saida, nome_arquivo)

            self.janela.after(0, self.atualizar_log, f"  Gerado: {os.path.basename(caminho_final)}", "padrao")

            progresso = (indice + 1) / len(linhas_validas)
            self.janela.after(0, lambda valor=progresso: self.barra_progresso.set(valor))


if __name__ == "__main__":
    raiz = ctk.CTk()
    app = AplicacaoPfoi(raiz)
    raiz.mainloop()
